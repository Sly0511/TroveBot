# Priority: 5
import asyncio
import json
import re
import typing
import urllib.request as urlget
from datetime import datetime, timedelta

import aiohttp
import discord
import html2text
import pytz
from bs4 import BeautifulSoup
from discord.ext import commands, tasks
from openpyxl import load_workbook
from pushbullet import Pushbullet

from utils.CustomObjects import CEmbed


class Tasks(commands.Cog):
    """Bot Tasks"""
    def __init__(self, bot):
        self.bot = bot
        self.pb = Pushbullet("o.b2ourOOHZtVViSoF8dXnZuFpZz2Zznsl")
        self.session = aiohttp.ClientSession()
        self.forums = [
            ["PTS Patches", "pts", "forumdisplay.php?50-Limited-Time-PTS-Discussion", "stickies"], # PTS Patch notes
            ["PTS Posts", "ptsposts", "forumdisplay.php?50-Limited-Time-PTS-Discussion", "threads"], # PTS User Posts
            ["Live Patches", "livepatches", "forumdisplay.php?5-Patch-Notes", "threads"], # Live Patch Notes
            ["Console Patches", "consolepatches", "forumdisplay.php?40-Console-Patch-Notes", "threads"], # Console Patch Notes
            #["Dev Tracker", "devtracker", "forumdisplay.php?3-Dev-Tracker", "threads"] # Dev Tracker
        ]
        print("Task Handler:\n|")
        self.sheet_timer.start()
        print("|--> Running Spreadsheet Retriever...")
        self.forum_retriever.start()
        self.reddit_retriever.start()
        print("|--> Running Forum Retriever...")
        self.forum_poster.start()
        self.reddit_poster.start()
        print("|--> Running Forum Poster...")
        self.bot_stats.start()
        self.clock.start()
        self.dragon_merchant_voice.start()
        self.dragon_merchant_text.start()
        print("|--> Running Clock...")
        self.daily_text_channels.start()
        self.daily_voice_channels.start()
        self.weekly_text_channels.start()
        self.weekly_voice_channels.start()
        print("|--> Running Automation tasks")

    def cog_unload(self):
        self.sheet_timer.cancel()
        self.forum_retriever.cancel()
        self.forum_poster.cancel()
        self.reddit_retriever.cancel()
        self.reddit_poster.cancel()
        self.bot_stats.cancel()
        self.clock.cancel()
        self.dragon_merchant_voice.cancel()
        self.dragon_merchant_text.cancel()
        self.daily_text_channels.cancel()
        self.daily_voice_channels.cancel()
        self.weekly_text_channels.cancel()
        self.weekly_voice_channels.cancel()
        asyncio.create_task(self.terminate_session())

    async def terminate_session(self):
        await self.session.close()

    async def wait_midnight(self):
        now = self.bot.Trove.time.now
        time = (23 - now.hour) * 3600 + (3540 - now.minute * 60) + (60 - now.second)
        await asyncio.sleep(time)

    @property
    def forums_time(self):
        return datetime.utcnow().astimezone(pytz.timezone("GMT")) - timedelta(hours=7)

    def convert_forum_time(self, time):
        if "today" in time.lower():
            new_time = self.forums_time
        elif "yesterday" in time.lower():
            new_time = self.forums_time - timedelta(days=1)
        else:
            return time
        return time.replace("Yesterday", new_time.strftime("%m-%d-%Y")).replace("Today", new_time.strftime("%m-%d-%Y"))

# Trove

    @tasks.loop(seconds=300)
    async def reddit_retriever(self):
        if self.bot.is_clone:
            return
        await self.bot.wait_until_ready()
        try:
            request = await self.bot.AIOSession.get("https://www.reddit.com/r/Trove/new.json")
            about = await self.bot.AIOSession.get("https://www.reddit.com/r/Trove/about.json")
            content = await request.text()
            about_content = await about.text()
            data = json.loads(content)
            about_data = json.loads(about_content)
        except:
            return
        posts = data["data"]["children"]
        for raw_post in posts:
            post = raw_post["data"]
            media = post["preview"]["images"] if "preview" in post and "images" in post["preview"] else None
            db_post = {
                "_id": post["id"],
                "title": post["title"],
                "description": post["selftext"],
                "author": post["author"],
                "image": media[0]["source"]["url"].replace("amp;", "") if media else None,
                "url": "https://www.reddit.com" + post["permalink"],
                "created_at": int(post["created_utc"]),
                "icon": about_data["data"]["community_icon"].split("?")[0],
                "type": "reddit"
            }
            try:
                await self.bot.db.db_forums.insert_one(db_post)
            except:
                ...

    @tasks.loop(seconds=180)
    async def reddit_poster(self):
        await self.bot.wait_until_ready()
        guilds = [g.id for g in self.bot.guilds]
        await self.bot.db.db_servers.update_many({f"forums_posts.reddit": {"$exists": False}}, {"$set": {f"forums_posts.reddit": {"channel": None, "posted": {}}}})
        guilds_settings = await self.bot.db.db_servers.find({"_id": {"$in": guilds}, f"forums_posts.reddit.channel": {"$ne": None}}, {"forums_posts": 1}).to_list(length=9999)
        forums_posts = await self.bot.db.db_forums.find({"type": "reddit"}).sort("created_at", 1).to_list(length=9999)
        for g in guilds_settings:
            guild = self.bot.get_guild(g["_id"])
            channel = guild.get_channel(g["forums_posts"]["reddit"]["channel"])
            if not channel:
                continue
            for forum_post in forums_posts:
                # if forum_post["created_at"] < channel.created_at.timestamp():
                #     continue
                e = CEmbed()
                e.title = forum_post["title"]
                e.url = forum_post["url"]
                e.description = forum_post["description"][:2041]
                e.color = discord.Color.random()
                e.set_author(name="r/Trove", url="https://www.reddit.com/r/Trove", icon_url=forum_post["icon"])
                e.timestamp = datetime.utcfromtimestamp(forum_post["created_at"])
                e.set_footer(text="Post by "+forum_post["author"])
                e.set_image(url=forum_post["image"])
                if len(e.description) == 2041:
                    e.description += "**...**"
                if forum_post["_id"] in g["forums_posts"]["reddit"]["posted"]:
                    continue
                try:
                    msg = await channel.send(embed=e)
                    try:
                        await msg.publish()
                    except:
                        pass
                except:
                    continue
                g["forums_posts"]["reddit"]["posted"][str(forum_post["_id"])] = [forum_post["created_at"], msg.id]
            await self.bot.db.db_servers.update_one({"_id": guild.id}, {"$set": {f"forums_posts.reddit.posted": g["forums_posts"]["reddit"]["posted"]}})

    @tasks.loop(seconds=180)
    async def forum_retriever(self):
        if self.bot.is_clone:
            return
        await self.bot.wait_until_ready()
        htmlhandler = html2text.HTML2Text()
        htmlhandler.ignore_images = True
        htmlhandler.ignore_emphasis = True
        htmlhandler.body_width = 0
        # Get page content
        for forum in self.forums:
            url_part = 'http://forums.trovegame.com/'
            url = f'{url_part}{forum[2]}'
            response = await (await self.session.get(url)).text()
            soup = BeautifulSoup(response, "html.parser")
            # Check database entries
            saved = await self.bot.db.db_forums.find({"type": forum[0]}).to_list(length=999999)
            saved_post_ids = [p["_id"] for p in saved]
            # Get all posts
            ct = soup.findAll("ol", id=forum[3])[0]
            ct = ct.find_all("li", id=lambda x: x and x.startswith('thread_'))
            posts = []
            i = 0
            for cta in ct:
                regex = r"([a-z0-9_]*) on ((?:[0-9]{2}-[0-9]{2}-[0-9]{4}|today|yesterday) [0-9]{2}:[0-9]{2} (?:am|pm))"
                raw_author = cta.find_all("a", {"class": lambda x: x and x.startswith('username')})[0].get("title")
                author_data = re.findall(regex, raw_author, re.IGNORECASE)[0]
                time = self.convert_forum_time(author_data[1])
                post_header = cta.find_all('a', {"class": "title"})[0]
                created_at = (datetime.strptime(time, "%m-%d-%Y %I:%M %p").replace(tzinfo=pytz.timezone("gmt"))+timedelta(hours=7)).astimezone(pytz.utc).timestamp()
                post = {
                    "_id": int(cta.get('id').replace("thread_", "")),
                    "author": author_data[0],
                    "type": forum[0],
                    "title": post_header.string,
                    "link": url_part + post_header.get("href").split("&")[0],
                    "created_at": created_at,
                    "edited_at": created_at
                }
                #if i == 0 and post["_id"] == 124849:
                #    print("PTS forums were cleared!")
                # Avoid default posts
                if post["_id"] in [106819, 124849]:
                    continue
                i += 1
                if datetime.utcnow().timestamp() - post["created_at"] > 86400*3:
                    continue
                response2 = await (await self.session.get(post["link"])).text()
                soup2 = BeautifulSoup(response2, "html.parser")
                htmlpost = soup2.findAll("div", {"class" : "content"})[0]
                post["html"] = str(htmlpost)
                text = htmlhandler.handle(str(htmlpost)).replace(" * ", "► ")
                regex = r"\[(.*)\]\((.*)\)"
                results = re.findall(regex, text)
                for result in results:
                    if not result[0]:
                        text = text.replace(f"![]({result[1]})", "")
                    else:
                        t = result[0].replace("\n", "") 
                        l = result[1].replace("\n", "")
                        text = text.replace(f"[{result[0]}]({result[1]})", f" [{t}]({l}) ")
                text = self.fixup_markdown_formatting(text)
                post["content"] = text
                if post["_id"] in saved_post_ids:
                    sp = list(filter(lambda x: x["_id"] == post["_id"], saved))[0]
                    if sp and sp == post:
                        continue
                    post["edited_at"] = int(datetime.utcnow().timestamp())
                posts.append(post)
                await asyncio.sleep(5)
            if posts:
                for post in posts:
                    await self.bot.db.db_forums.update_one({"_id": post["_id"]}, {"$set": post}, upsert=True)
        # except Exception as e:
        #     print(f"{e.__traceback__.tb_lineno} - {e}")
        #     print(f"Forum Retriever task failed! -> {forum[0]}")
        #     pass

    def fixup_markdown_formatting(self, text):
        # Strip off table formatting
        text = re.sub(r'(^|\n)\|\s*', r'\1', text)
        # Strip off extra emphasis
        text = re.sub(r'\*\*', '', text)
        # Remove trailing whitespace and leading newlines
        text = re.sub(r' *$', '', text)
        text = re.sub(r'\n\n+', r'\n\n', text)
        text = re.sub(r'^\n+', '', text)
        return text

    @commands.command(slash_command=True, slash_command_guilds=[834505270075457627], help="Import posts from forums onto bot", hidden=True)
    async def import_posts(self, ctx,
        url=commands.Option(description="URL of forum page"),
        posttype: typing.Literal["PTS Patches", "PTS Posts", "Live Patches", "Console Patches"]=commands.Option(description="Select a type for the posts to be saved as."),
        stickies: typing.Literal[1, 0]=commands.Option(description="Find in stickies?")):
        if ctx.author.id != 565097923025567755:
            return await ctx.send("Fuck off", ephemeral=True)
        await ctx.defer(trigger_typing=True)
        url_part = 'http://forums.trovegame.com/'
        forum = [posttype, 0, 0]
        if int(stickies):
            forum.append("stickies")
        else:
            forum.append("threads")
        htmlhandler = html2text.HTML2Text()
        htmlhandler.ignore_images = True
        htmlhandler.ignore_emphasis = True
        htmlhandler.body_width = 0
        response = await (await self.session.get(url)).text()
        soup = BeautifulSoup(response, "html.parser")
        # Check database entries
        saved = await self.bot.db.db_forums.find({"type": forum[0]}).to_list(length=999999)
        saved_post_ids = [p["_id"] for p in saved]
        # Get all posts
        ct = soup.findAll("ol", id=forum[3])[0]
        ct = ct.find_all("li", id=lambda x: x and x.startswith('thread_'))
        posts = []
        i = 0
        for cta in ct:
            regex = r"([a-z0-9_]*) on ((?:[0-9]{2}-[0-9]{2}-[0-9]{4}|today|yesterday) [0-9]{2}:[0-9]{2} (?:am|pm))"
            raw_author = cta.find_all("a", {"class": lambda x: x and x.startswith('username')})[0].get("title")
            author_data = re.findall(regex, raw_author, re.IGNORECASE)[0]
            time = self.convert_forum_time(author_data[1])
            post_header = cta.find_all('a', {"class": "title"})[0]
            created_at = (datetime.strptime(time, "%m-%d-%Y %I:%M %p").replace(tzinfo=pytz.timezone("gmt"))+timedelta(hours=7)).astimezone(pytz.utc).timestamp()
            post = {
                "_id": int(cta.get('id').replace("thread_", "")),
                "author": author_data[0],
                "type": forum[0],
                "title": post_header.string,
                "link": url_part + post_header.get("href").split("&")[0],
                "created_at": created_at,
                "edited_at": created_at
            }
            #if i == 0 and post["_id"] == 124849:
            #    print("PTS forums were cleared!")
            # Avoid default posts
            if post["_id"] in [106819, 124849]:
                continue
            i += 1
            # if datetime.utcnow().timestamp() - post["created_at"] > 86400*3:
            #     print("lol")
            #     continue
            response2 = await (await self.session.get(post["link"])).text()
            soup2 = BeautifulSoup(response2, "html.parser")
            htmlpost = soup2.findAll("div", {"class" : "content"})[0]
            post["html"] = str(htmlpost)
            text = htmlhandler.handle(str(htmlpost)).replace(" * ", "► ")
            regex = r"\[(.*)\]\((.*)\)"
            results = re.findall(regex, text)
            for result in results:
                if not result[0]:
                    text = text.replace(f"![]({result[1]})", "")
                else:
                    t = result[0].replace("\n", "") 
                    l = result[1].replace("\n", "")
                    text = text.replace(f"[{result[0]}]({result[1]})", f" [{t}]({l}) ")
            text = self.fixup_markdown_formatting(text)
            post["content"] = text
            if post["_id"] in saved_post_ids:
                continue
                sp = list(filter(lambda x: x["_id"] == post["_id"], saved))[0]
                if sp and sp == post:
                    continue
                post["edited_at"] = int(datetime.utcnow().timestamp())
            posts.append(post)
        if posts:
            for post in posts:
                await self.bot.db.db_forums.update_one({"_id": post["_id"]}, {"$set": post}, upsert=True)
            return await ctx.send(f"Done, Imported {len(posts)}")
        await ctx.send("Done, No imported posts")

    @tasks.loop(seconds=180)
    async def forum_poster(self):
        await self.bot.wait_until_ready()
        guilds = [g.id for g in self.bot.guilds]
        for post_type in self.forums:
            await self.bot.db.db_servers.update_many({f"forums_posts.{post_type[1]}": {"$exists": False}}, {"$set": {f"forums_posts.{post_type[1]}": {"channel": None, "posted": {}}}})
            guilds_settings = await self.bot.db.db_servers.find({"_id": {"$in": guilds}, f"forums_posts.{post_type[1]}.channel": {"$ne": None}}, {"forums_posts": 1}).to_list(length=9999)
            if not len(guilds_settings):
                continue
            forums_posts = await self.bot.db.db_forums.find({"type": post_type[0]}, {"html": 0}).sort("created_at", 1).to_list(length=9999)
            for g in guilds_settings:
                guild = self.bot.get_guild(g["_id"])
                channel = guild.get_channel(g["forums_posts"][post_type[1]]["channel"])
                if not channel:
                    continue
                for forum_post in forums_posts:
                    e = CEmbed()
                    e.color = discord.Color.random()
                    e.set_author(name=forum_post["author"])
                    e.timestamp = datetime.utcfromtimestamp(forum_post["created_at"])
                    if forum_post["created_at"] < channel.created_at.timestamp():
                        continue
                    if datetime.utcnow().timestamp() - forum_post["created_at"] > 604800: 
                        continue
                    e.title = forum_post["title"]
                    e.url = forum_post["link"]
                    e.set_footer(text=forum_post["_id"])
                    add = f"[Dark Mode Page](https://trove.slynx.xyz/posts/{forum_post['_id']})"
                    e.description = forum_post["content"].strip() + f"\n\n{add}"
                    if len(e.description) > 2048:
                        e.description = f"Post is too big to show in discord, check it out at {add}"
                    if str(forum_post["_id"]) in g["forums_posts"][post_type[1]]["posted"].keys():
                        if forum_post["edited_at"] == g["forums_posts"][post_type[1]]["posted"][str(forum_post["_id"])][0]:
                            continue
                        e.footer.text += " | This post was edited at"
                        e.timestamp = datetime.utcfromtimestamp(forum_post["edited_at"])
                        try:
                            msg = await channel.fetch_message(g["forums_posts"][post_type[1]]["posted"][str(forum_post["_id"])][1])
                        except:
                            continue
                        try:
                            await msg.edit(embed=e)
                            g["forums_posts"][post_type[1]]["posted"][str(forum_post["_id"])][0] = forum_post["edited_at"]
                            continue
                        except:
                            pass
                    try:
                        msg = await channel.send(embed=e)
                        try:
                            await msg.publish()
                        except:
                            pass
                    except:
                        continue
                    if guild.id == 567505514108289044 and forum_post["type"] == "PTS Patches":
                        for device in self.pb.devices:
                            if device.device_iden != "ujvgNES5wOasjxJcPjX6Qe":
                                continue
                            device.push_note(forum_post["title"], forum_post["content"])
                    g["forums_posts"][post_type[1]]["posted"][str(forum_post["_id"])] = [forum_post["edited_at"], msg.id]
                await self.bot.db.db_servers.update_one({"_id": guild.id}, {"$set": {f"forums_posts.{post_type[1]}.posted": g["forums_posts"][post_type[1]]["posted"]}})

    @tasks.loop(seconds=5)
    async def bot_stats(self):
        if self.bot.is_clone:
            return
        await self.bot.wait_until_ready()
        time = self.bot.Trove.time.now
        if time.minute%10:
            return
        channels = [
            {
                "id": 864087320978391041,
                "text": f"Users: {len(self.bot.users)}"
            },
            {
                "id": 864087351709794304,
                "text": f"Servers: {len(self.bot.guilds)}"
            }
        ]
        for ch in channels:
            channel = self.bot.get_channel(ch["id"])
            if not ch:
                continue
            if channel.name == ch["text"]:
                continue
            try:
                await channel.edit(name=ch["text"])
            except:
                ...
        await asyncio.sleep(60)

    @tasks.loop(seconds=5)
    async def clock(self):
        time = self.bot.Trove.time.now
        data = await self.bot.db.db_servers.find({"clock.channel": {"$ne": None}}, {"clock": 1}).to_list(length=99999)
        if not data:
            return
        for server in data:
            timeout = server["clock"]["slowmode"]
            if timeout:
                if time.minute%timeout:
                    continue
            else:
                if time.minute != timeout:
                    continue
            channel = self.bot.get_channel(server["clock"]["channel"])
            if not channel:
                continue
            _format = server["clock"]["format"] or "⌛{hour24}:{minute}"
            text = _format.replace(r"{minute}", f"{time.minute:02d}")
            text = text.replace(r"{hour24}", f"{time.hour:02d}")
            text = text.replace(r"{hour12}", f"{time.hour if time.hour <= 12 else time.hour-12:02d}")
            text = text.replace(r"{daysplit}", "AM" if time.hour < 12 else "PM")
            if channel.name == text:
                continue
            try:
                asyncio.create_task(channel.edit(name=text))
            except:
                ...
        await asyncio.sleep(15)

    @tasks.loop(seconds=60)
    async def sheet_timer(self):
        await self.bot.wait_until_ready()
        def load_sheets():
            if not self.bot.is_clone:
                urlget.urlretrieve("https://docs.google.com/spreadsheets/d/1hsz9Xhf52xjX0pcfb95Tm3-MvVh9lMfq4bGi5fo7zCs/export?format=xlsx&id=1hsz9Xhf52xjX0pcfb95Tm3-MvVh9lMfq4bGi5fo7zCs", 'data/sheets/luxion_sheet.xlsx')
                urlget.urlretrieve("https://docs.google.com/spreadsheets/d/1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8/export?format=xlsx&id=1YBf3__CPCy9iL4HDEoF1_q88vaFtAR6mA4GZxIzOEG8", 'data/sheets/memento_list_sheet.xlsx')
            try:
                self.bot.Trove.sheets["summer"] = load_workbook(filename="data/sheets/luxion_sheet.xlsx")
            except:
                pass
            self.bot.Trove.sheets["memento_list"] = load_workbook(filename="data/sheets/memento_list_sheet.xlsx", data_only=True)
        await self.bot.loop.run_in_executor(None, load_sheets)

    @tasks.loop(seconds=5)
    async def dragon_merchant_voice(self):
        data = await self.bot.db.db_servers.find({"automation.dragon_merchant.voice.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        text = "🐲"
        if self.bot.Trove.time.is_luxion:
            text += "Luxion in Hub"
        elif self.bot.Trove.time.is_corruxion:
            text += "Corruxion in Hub"
        else:
            text += "No Dragon Merchant"
        if not data:
            return
        for server in data:
            channel = self.bot.get_channel(server["automation"]["dragon_merchant"]["voice"]["channel"])
            if not channel:
                continue
            if channel.name == text:
                continue
            try:
                await channel.edit(name=text)
            except:
                ...
        await self.wait_midnight()

    @tasks.loop(seconds=5)
    async def dragon_merchant_text(self):
        now = self.bot.Trove.time.now
        if not (now.hour == 0 and now.minute == 0):
            return
        data = await self.bot.db.db_servers.find({"automation.dragon_merchant.text.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        if not data:
            return
        e = CEmbed()
        if self.bot.Trove.time.luxion_start.day == now.day:
            dragon = "Luxion"
            avatar = "https://i.imgur.com/9eOV0JD.png"
            e.color = 0xc8ad18
            e.set_image(url="https://i.imgur.com/zWkZ9Xd.png")
        elif self.bot.Trove.time.corruxion_start.day == now.day:
            dragon = "Corruxion"
            avatar = "https://i.imgur.com/BPNdE1w.png"
            e.color = 0x850eff
            e.set_image(url="https://i.imgur.com/AqXZGaJ.png")
        else:
            return
        e.set_author(name="Dragon Merchant", icon_url=avatar)
        e.title = f"{dragon} has landed in the Hub"
        for server in data:
            channel = self.bot.get_channel(server["automation"]["dragon_merchant"]["text"]["channel"])
            if channel:
                role = self.bot.get_guild(server["_id"]).get_role(server["automation"]["dragon_merchant"]["text"]["role"])
                content = role.mention if role else None
                msg = await channel.send(content=content, embed=e)
                try:
                    await msg.publish()
                except:
                    pass
        await asyncio.sleep(60)

    @tasks.loop(seconds=5)
    async def daily_text_channels(self):
        if not (self.bot.Trove.time.now.hour == 0 and self.bot.Trove.time.now.minute == 0):
            return
        data = await self.bot.db.db_servers.find({"automation.daily.text.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        if not data:
            return
        weekday = self.bot.Trove.daily_data[str(self.bot.Trove.time.now.weekday())]
        e = CEmbed()
        e.color = discord.Color(int(weekday["color"], 16))
        e.set_author(name=weekday["name"], icon_url=self.bot.user.avatar)
        e.title = "Server Daily Reset"
        e.description = "<:NoPatron:845646875150778418> **Regular Buffs**```\n" + "\n".join(weekday["normal_buffs"]) + "\n```"
        e.description += "\n<:Patron:845646874752057365> **Patron Buffs**```\n" + "\n".join(weekday["premium_buffs"]) + "\n```"
        e.set_image(url=weekday["banner"])
        for server in data:
            channel = self.bot.get_channel(server["automation"]["daily"]["text"]["channel"])
            if channel:
                role = self.bot.get_guild(server["_id"]).get_role(server["automation"]["daily"]["text"]["role"])
                content = role.mention if role else None
                msg = await channel.send(content=content, embed=e)
                try:
                    await msg.publish()
                except:
                    pass
        await asyncio.sleep(60)

    @tasks.loop(seconds=5)
    async def daily_voice_channels(self):
        data = await self.bot.db.db_servers.find({"automation.daily.voice.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        if not data:
            return
        for server in data:
            dailies = ["🏹Delve Day", "🐟Gathering Day", "💎Gem Day", "🌀Adventure Day", "🐉Dragon Day", "💉XP Day", "💰Loot Day"]
            channel = self.bot.get_channel(server["automation"]["daily"]["voice"]["channel"])
            daily = dailies[self.bot.Trove.time.now.weekday()]
            if channel and channel.name != daily:
                try:
                    await channel.edit(name=daily)
                except:
                    ...
        await self.wait_midnight()

    @tasks.loop(seconds=5)
    async def weekly_text_channels(self):
        if not (self.bot.Trove.time.now.hour == 0 and self.bot.Trove.time.now.minute == 0 and self.bot.Trove.time.now.weekday() == 0):
            return
        data = await self.bot.db.db_servers.find({"automation.weekly.text.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        if not data:
            return
        weekly = self.bot.Trove.weekly_data[str(self.bot.Trove.time.weekly_time)]
        e = CEmbed()
        e.color = discord.Color(int(weekly["color"], 16))
        e.set_author(name=weekly["name"], icon_url=self.bot.user.avatar)
        e.title = "Server Weekly Reset"
        e.description = "**Weekly Bonus**```\n" + "\n".join(weekly["buffs"]) + "\n```"
        e.set_image(url=weekly["banner"])
        for server in data:
            channel = self.bot.get_channel(server["automation"]["weekly"]["text"]["channel"])
            if channel:
                role = self.bot.get_guild(server["_id"]).get_role(server["automation"]["weekly"]["text"]["role"])
                content = role.mention if role else None
                msg = await channel.send(content=content, embed=e)
                try:
                    await msg.publish()
                except:
                    pass
        await asyncio.sleep(60)

    @tasks.loop(seconds=5)
    async def weekly_voice_channels(self):
        data = await self.bot.db.db_servers.find({"automation.weekly.voice.channel": {"$ne": None}}, {"automation": 1}).to_list(length=99999)
        if data:
            weekly = self.bot.Trove.weekly_data[str(self.bot.Trove.time.weekly_time)]
            buff = f"{weekly['emoji']}{weekly['name']}"
            for server in data:
                channel = self.bot.get_channel(server["automation"]["weekly"]["voice"]["channel"])
                if channel and channel.name != buff:
                    try:
                        await channel.edit(name=buff)
                    except:
                        ...
        await self.wait_midnight()

def setup(bot):
    n = Tasks(bot)
    bot.add_cog(n)
